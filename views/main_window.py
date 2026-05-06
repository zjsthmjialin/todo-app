"""
主窗口 - 整合所有组件
"""

from PySide6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                               QPushButton, QLabel, QLineEdit, QScrollArea,
                               QStatusBar, QMessageBox, QDateEdit, QTimeEdit,
                               QDialog, QTextEdit, QCheckBox, QSizePolicy,
                               QGraphicsDropShadowEffect, QMenu, QComboBox,
                               QSystemTrayIcon, QFileDialog, QProgressBar)
from PySide6.QtCore import Qt, QDate, QTime, Signal, QMimeData, QTimer, QRect, QEvent
from PySide6.QtGui import QDrag, QEnterEvent, QColor, QAction, QCursor, QIcon, QFont, QPixmap, QPainter, QPainterPath
import sys
import os
import json
from datetime import datetime

from controllers import TaskController
from utils import ThemeManager


class TaskCard(QWidget):
    """任务卡片组件"""
    delete_signal = Signal(int)
    toggle_signal = Signal(int)
    edit_signal = Signal(int)

    def __init__(self, task, theme_manager, parent=None):
        super().__init__(parent)
        self.task = task
        self.theme_manager = theme_manager
        self.is_expanded = False
        self.is_dragging = False
        self.setup_ui()

    def setup_ui(self):
        self.setObjectName("taskCard")
        self.setMinimumHeight(48)

        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 第一行：卡片主要内容
        card_layout = QHBoxLayout()
        card_layout.setContentsMargins(8, 6, 8, 6)
        card_layout.setSpacing(8)

        # 勾选框
        self.checkbox = QPushButton()
        self.checkbox.setFixedSize(22, 22)
        self.checkbox.setObjectName("taskCheckbox")
        self.checkbox.setCheckable(True)
        self.checkbox.setChecked(self.task.is_completed)
        self.checkbox.clicked.connect(lambda: self.toggle_signal.emit(self.task.id))
        card_layout.addWidget(self.checkbox)

        # 任务内容
        content_layout = QVBoxLayout()
        content_layout.setSpacing(2)

        title_layout = QHBoxLayout()
        if self.task.is_important:
            self.star = QLabel("⭐")
            self.star.setObjectName("starLabel")
            title_layout.addWidget(self.star)

        self.title_label = QLabel(self.task.title)
        self.title_label.setObjectName("taskTitle")
        if self.task.is_completed:
            self.title_label.setStyleSheet("text-decoration: line-through; color: #A19F9D;")
        title_layout.addWidget(self.title_label)

        # 显示创建日期
        if self.task.created_at:
            from datetime import datetime
            try:
                created_date = datetime.strptime(self.task.created_at, "%Y-%m-%d %H:%M:%S")
                date_str = created_date.strftime("%Y-%m-%d")
                self.created_label = QLabel(date_str)
                self.created_label.setObjectName("createdLabel")
                title_layout.addWidget(self.created_label)
            except:
                pass

        title_layout.addStretch()

        content_layout.addLayout(title_layout)

        if self.task.due_date:
            date_text = self.task.due_date
            if self.task.due_time:
                date_text += f" {self.task.due_time}"
            self.date_label = QLabel(date_text)
            self.date_label.setObjectName("taskDate")
            content_layout.addWidget(self.date_label)

        card_layout.addLayout(content_layout, 1)

        # 编辑按钮
        self.edit_btn = QPushButton("✎")
        self.edit_btn.setFixedSize(28, 28)
        self.edit_btn.setObjectName("editButton")
        self.edit_btn.clicked.connect(self.emit_edit)
        card_layout.addWidget(self.edit_btn)

        # 删除按钮
        self.delete_btn = QPushButton("×")
        self.delete_btn.setFixedSize(28, 28)
        self.delete_btn.setObjectName("deleteButton")
        self.delete_btn.clicked.connect(lambda: self.delete_signal.emit(self.task.id))
        card_layout.addWidget(self.delete_btn)

        main_layout.addLayout(card_layout)

        # 详情区域（默认隐藏）
        self.detail_widget = QWidget()
        self.detail_widget.setObjectName("taskDetail")
        detail_layout = QVBoxLayout(self.detail_widget)
        detail_layout.setContentsMargins(38, 0, 20, 8)
        detail_layout.setSpacing(4)

        if self.task.description:
            desc_label = QLabel(self.task.description)
            desc_label.setObjectName("taskDescription")
            desc_label.setWordWrap(True)
            detail_layout.addWidget(desc_label)

        self.detail_widget.hide()
        main_layout.addWidget(self.detail_widget)

        # 应用主题
        self.update_style()

    def update_style(self):
        theme = self.theme_manager.get_current_theme()
        primary = theme["primary"]
        self.checkbox.setStyleSheet(f"""
            QPushButton#taskCheckbox {{
                border: 1.5px solid #c8bfb0;
                border-radius: 6px;
                background: white;
                min-width: 22px;
                min-height: 22px;
            }}
            QPushButton#taskCheckbox:checked {{
                background-color: {primary};
                border-color: {primary};
            }}
        """)

    def emit_edit(self):
        self.window().edit_task(self.task)

    def toggle_expand(self):
        """切换详情展开/收起"""
        if hasattr(self, 'detail_widget'):
            if self.is_expanded:
                self.detail_widget.hide()
                self.is_expanded = False
            else:
                if self.task.description:
                    self.detail_widget.show()
                    self.is_expanded = True

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            # 检查是否点击在按钮上
            widget = self.childAt(event.pos())
            if widget and (widget == self.edit_btn or widget == self.delete_btn or widget == self.checkbox):
                return

            # 左键点击：显示/隐藏描述
            self.toggle_expand()
            event.accept()

        elif event.button() == Qt.RightButton:
            # 右键按下：启动拖拽
            self.is_dragging = True
            self.drag_start_position = event.pos()
            self.drag_task_id = self.task.id

    def mouseMoveEvent(self, event):
        # 只处理右键拖拽
        if hasattr(self, 'is_dragging') and self.is_dragging:
            if (event.buttons() & Qt.RightButton):
                if (event.pos() - self.drag_start_position).manhattanLength() > 10:
                    mime_data = QMimeData()
                    mime_data.setText(str(self.drag_task_id))
                    drag = QDrag(self)
                    drag.setMimeData(mime_data)
                    drag.exec(Qt.CopyAction)
                    self.is_dragging = False


class TaskEditDialog(QDialog):
    """任务编辑对话框"""

    def __init__(self, task=None, parent=None):
        super().__init__(parent)
        self.task = task
        self.setWindowTitle("编辑任务" if task else "添加任务")
        self.setMinimumWidth(400)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)

        # 标题
        title_layout = QHBoxLayout()
        title_layout.addWidget(QLabel("任务标题:"))
        self.title_input = QLineEdit()
        if self.task:
            self.title_input.setText(self.task.title)
        title_layout.addWidget(self.title_input)
        layout.addLayout(title_layout)

        # 描述
        desc_layout = QHBoxLayout()
        desc_layout.addWidget(QLabel("描述:"))
        self.desc_input = QTextEdit()
        self.desc_input.setMaximumHeight(80)
        if self.task:
            self.desc_input.setText(self.task.description)
        desc_layout.addWidget(self.desc_input)
        layout.addLayout(desc_layout)

        # 截止日期
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("截止日期:"))
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        if self.task and self.task.due_date:
            self.date_edit.setDate(QDate.fromString(self.task.due_date, "yyyy-MM-dd"))
        date_layout.addWidget(self.date_edit)
        layout.addLayout(date_layout)

        # 截止时间
        time_layout = QHBoxLayout()
        time_layout.addWidget(QLabel("截止时间:"))
        self.time_edit = QTimeEdit()
        self.time_edit.setDisplayFormat("HH:mm")
        if self.task and self.task.due_time:
            self.time_edit.setTime(QTime.fromString(self.task.due_time, "HH:mm"))
        time_layout.addWidget(self.time_edit)
        layout.addLayout(time_layout)

        # 重要标记
        self.important_check = QCheckBox("⭐ 标记为重要")
        if self.task:
            self.important_check.setChecked(self.task.is_important)
        layout.addWidget(self.important_check)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.save_btn)

        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)

        layout.addLayout(btn_layout)

    def get_task_data(self) -> dict:
        return {
            'title': self.title_input.text(),
            'description': self.desc_input.toPlainText(),
            'due_date': self.date_edit.date().toString("yyyy-MM-dd"),
            'due_time': self.time_edit.time().toString("HH:mm"),
            'is_important': self.important_check.isChecked(),
        }


class MainWindow(QMainWindow):
    """主窗口"""

    def __init__(self):
        super().__init__()
        self.controller = TaskController()
        self.theme_manager = ThemeManager(self.controller.db)

        self.current_list_id = 0
        self.current_list_name = "我的日程"
        self.sidebar_visible = True
        self.current_filter = "all"  # all, active, completed

        self.setWindowTitle("JinSongToDo")
        self.setMinimumSize(700, 500)

        # 创建菜单栏
        self.create_menu_bar()

        # 设置窗口图标
        import os
        icon_path = os.path.join(os.path.dirname(__file__), '..', 'todo-icon-final_3.png')
        icon_path = os.path.abspath(icon_path)
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setup_ui()
        self.apply_theme()

        # 创建系统托盘（在UI之后）
        self.create_tray_icon()
        self.load_lists()
        self.on_list_selected(1, "我的日程")  # 默认选中"我的日程"
        self.load_tasks()

    def create_tray_icon(self):
        """创建系统托盘图标"""
        import os

        # 创建托盘图标
        self.tray_icon = QSystemTrayIcon(self)

        # 设置图标
        icon_path = os.path.join(os.path.dirname(__file__), '..', 'todo-icon-final_3.png')
        icon_path = os.path.abspath(icon_path)

        if os.path.exists(icon_path):
            self.tray_icon.setIcon(QIcon(icon_path))
            self.tray_icon.setVisible(True)
        else:
            print(f"Icon not found at: {icon_path}")

        # 托盘气泡提示
        self.tray_icon.setToolTip("JinSongToDo - 点击恢复窗口")

        # 创建托盘菜单
        tray_menu = QMenu(self)

        show_action = QAction("显示窗口", self)
        show_action.triggered.connect(self.show_window)
        tray_menu.addAction(show_action)

        tray_menu.addSeparator()

        exit_action = QAction("退出", self)
        exit_action.triggered.connect(self.exit_application)
        tray_menu.addAction(exit_action)

        self.tray_icon.setContextMenu(tray_menu)

        # 点击托盘图标显示窗口
        self.tray_icon.activated.connect(self.on_tray_activated)

        # 托盘图标显示
        self.tray_icon.show()

    def on_tray_activated(self, reason):
        """托盘图标被激活"""
        if reason == QSystemTrayIcon.Trigger or reason == QSystemTrayIcon.DoubleClick:
            self.show_window()

    def show_window(self):
        """显示主窗口"""
        self.show()
        self.raise_()
        self.activateWindow()

    def exit_application(self):
        """退出应用程序"""
        self.tray_icon.hide()
        sys.exit(0)

    def create_menu_bar(self):
        """创建菜单栏"""
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu("文件")

        new_task_action = QAction("新建任务", self)
        new_task_action.setShortcut("Ctrl+N")
        new_task_action.triggered.connect(self.add_task)
        file_menu.addAction(new_task_action)

        file_menu.addSeparator()

        export_action = QAction("导出数据", self)
        export_action.triggered.connect(self.export_data)
        file_menu.addAction(export_action)

        import_action = QAction("导入数据", self)
        import_action.triggered.connect(self.import_data)
        file_menu.addAction(import_action)

        file_menu.addSeparator()

        exit_action = QAction("退出", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.exit_application)
        file_menu.addAction(exit_action)

        # 编辑菜单
        edit_menu = menubar.addMenu("编辑")

        select_all_action = QAction("全选", self)
        select_all_action.setShortcut("Ctrl+A")
        select_all_action.triggered.connect(self.select_all_tasks)
        edit_menu.addAction(select_all_action)

        delete_completed_action = QAction("删除已完成任务", self)
        delete_completed_action.triggered.connect(self.delete_completed_tasks)
        edit_menu.addAction(delete_completed_action)

        # 视图菜单
        view_menu = menubar.addMenu("视图")

        self.toggle_sidebar_action = QAction("隐藏侧边栏", self)
        self.toggle_sidebar_action.setShortcut("Ctrl+L")
        self.toggle_sidebar_action.triggered.connect(self.toggle_sidebar)
        view_menu.addAction(self.toggle_sidebar_action)

        # 主题颜色子菜单
        theme_menu = view_menu.addMenu("主题颜色")

        themes = self.theme_manager.get_all_themes()
        theme_keys = list(themes.keys())

        for key in theme_keys:
            theme = themes[key]
            color_action = QAction("", self)
            color_action.setToolTip(theme['name'])
            # 创建彩色的 ■ 图标
            pixmap = QPixmap(20, 20)
            pixmap.fill(QColor(theme['primary']))
            color_action.setIcon(QIcon(pixmap))
            color_action.triggered.connect(lambda checked, k=key: self.change_theme(k))
            theme_menu.addAction(color_action)

        # 帮助菜单
        help_menu = menubar.addMenu("帮助")

        about_action = QAction("关于", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

        shortcuts_action = QAction("快捷键说明", self)
        shortcuts_action.triggered.connect(self.show_shortcuts)
        help_menu.addAction(shortcuts_action)

    def toggle_pin(self, checked):
        """切换窗口置顶状态"""
        if checked:
            new_flags = self.windowFlags() | Qt.WindowStaysOnTopHint
            self.pin_btn.setStyleSheet("background-color: #f0ebe0; border-radius: 6px;")
            self.pin_btn.setIcon(create_star_icon("#a09080"))
        else:
            new_flags = self.windowFlags() & ~Qt.WindowStaysOnTopHint
            self.pin_btn.setStyleSheet("background-color: rgba(90,143,110,0.2); border-radius: 6px;")
            self.pin_btn.setIcon(create_star_icon("#5a8f6e"))

        # 使用低级别方法设置标志避免闪烁
        from ctypes import windll
        hwnd = int(self.winId())
        GWL_EXSTYLE = -20
        WS_EX_TOPMOST = 0x00000008

        if checked:
            windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE) | WS_EX_TOPMOST)
        else:
            windll.user32.SetWindowLongW(hwnd, GWL_EXSTYLE, windll.user32.GetWindowLongW(hwnd, GWL_EXSTYLE) & ~WS_EX_TOPMOST)

    def toggle_sidebar(self):
        """显示/隐藏侧边栏"""
        if self.sidebar_visible:
            self.sidebar.hide()
            self.sidebar_visible = False
            self.toggle_sidebar_action.setText("显示侧边栏")
        else:
            self.sidebar.show()
            self.sidebar_visible = True
            self.toggle_sidebar_action.setText("隐藏侧边栏")

    def export_data(self):
        """导出数据"""
        # 获取所有列表和任务
        lists = self.controller.get_all_lists()
        all_tasks = self.controller.get_tasks()

        # 格式选择对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("导出数据")
        dialog.setFixedSize(300, 120)

        layout = QVBoxLayout(dialog)

        # 格式选择
        format_layout = QHBoxLayout()
        format_layout.addWidget(QLabel("导出格式:"))
        format_combo = QComboBox()
        format_combo.addItems(["Markdown (.md)", "Excel (.xlsx)", "JSON (.json)"])
        format_layout.addWidget(format_combo)
        layout.addLayout(format_layout)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        export_btn = QPushButton("导出")
        export_btn.clicked.connect(lambda: self.do_export(dialog, format_combo.currentIndex(), lists, all_tasks))
        btn_layout.addWidget(export_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        dialog.exec()

    def do_export(self, dialog, format_index, lists, all_tasks):
        dialog.accept()
        if format_index == 0:
            self.export_markdown(lists, all_tasks)
        elif format_index == 1:
            self.export_excel(lists, all_tasks)
        elif format_index == 2:
            self.export_json(lists, all_tasks)

    def export_markdown(self, lists, tasks):
        """导出为 Markdown 格式"""
        file_path, _ = QFileDialog.getSaveFileName(self, "导出 Markdown", "",
                                                    "Markdown Files (*.md)")
        if not file_path:
            return
        if not file_path.endswith('.md'):
            file_path += '.md'

        md_content = "# 待办事项导出\n\n"
        md_content += f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

        # 按列表分组
        for lst in lists:
            list_tasks = [t for t in tasks if t.list_id == lst.id]
            if list_tasks:
                md_content += f"## {lst.icon} {lst.name}\n\n"
                for task in list_tasks:
                    checkbox = "[x]" if task.is_completed else "[ ]"
                    important = "⭐ " if task.is_important else ""
                    due = f" (截止: {task.due_date} {task.due_time or ''})" if task.due_date else ""
                    md_content += f"- {checkbox} {important}{task.title}{due}\n"
                    if task.description:
                        md_content += f"  - 描述: {task.description}\n"
                md_content += "\n"

        # 未分类任务
        uncategorized = [t for t in tasks if t.list_id == 0]
        if uncategorized:
            md_content += f"## 📋 未分类\n\n"
            for task in uncategorized:
                checkbox = "[x]" if task.is_completed else "[ ]"
                important = "⭐ " if task.is_important else ""
                md_content += f"- {checkbox} {important}{task.title}\n"

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(md_content)

        QMessageBox.information(self, "导出成功", f"已导出为 Markdown 格式:\n{file_path}")

    def export_excel(self, lists, tasks):
        """导出为 Excel 格式"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "导出失败", "请安装 openpyxl 库: pip install openpyxl")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "导出 Excel", "",
                                                    "Excel Files (*.xlsx)")
        if not file_path:
            return
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "待办事项"

        # 表头
        headers = ["列表", "任务", "描述", "截止日期", "截止时间", "重要", "已完成", "创建时间"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据
        task_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for row, task in enumerate(tasks, 2):
            list_name = next((lst.name for lst in lists if lst.id == task.list_id), "未分类")
            ws.cell(row=row, column=1, value=list_name)
            ws.cell(row=row, column=2, value=task.title)
            ws.cell(row=row, column=3, value=task.description or "")
            ws.cell(row=row, column=4, value=task.due_date or "")
            ws.cell(row=row, column=5, value=task.due_time or "")
            ws.cell(row=row, column=6, value="⭐" if task.is_important else "")
            ws.cell(row=row, column=7, value="是" if task.is_completed else "否")
            ws.cell(row=row, column=8, value=task.created_at or "")

            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = task_fill

        # 自动列宽
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 15

        wb.save(file_path)
        QMessageBox.information(self, "导出成功", f"已导出为 Excel 格式:\n{file_path}")

    def export_json(self, lists, tasks):
        """导出为 JSON 格式"""
        file_path, _ = QFileDialog.getSaveFileName(self, "导出 JSON", "",
                                                    "JSON Files (*.json)")
        if not file_path:
            return
        if not file_path.endswith('.json'):
            file_path += '.json'

        data = {
            "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "version": "1.0",
            "lists": [
                {
                    "id": lst.id,
                    "name": lst.name,
                    "icon": lst.icon,
                    "color": lst.color,
                    "is_system": lst.is_system
                } for lst in lists
            ],
            "tasks": [
                {
                    "id": t.id,
                    "title": t.title,
                    "description": t.description,
                    "is_important": t.is_important,
                    "is_completed": t.is_completed,
                    "due_date": t.due_date,
                    "due_time": t.due_time,
                    "list_id": t.list_id,
                    "created_at": t.created_at
                } for t in tasks
            ]
        }

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        QMessageBox.information(self, "导出成功", f"已导出为 JSON 格式:\n{file_path}")

    def import_data(self):
        """导入数据"""
        file_path, _ = QFileDialog.getOpenFileName(self, "导入 JSON", "",
                                                   "JSON Files (*.json)")
        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # 验证数据格式
            if 'tasks' not in data:
                QMessageBox.warning(self, "导入失败", "无效的数据格式")
                return

            # 导入任务
            imported_count = 0
            for task_data in data['tasks']:
                task = Task(
                    title=task_data.get('title', ''),
                    description=task_data.get('description', ''),
                    is_important=task_data.get('is_important', False),
                    is_completed=task_data.get('is_completed', False),
                    due_date=task_data.get('due_date'),
                    due_time=task_data.get('due_time'),
                    list_id=task_data.get('list_id', 0),
                )
                self.controller.db.create_task(task)
                imported_count += 1

            self.load_tasks()
            self.load_lists()
            QMessageBox.information(self, "导入成功", f"成功导入 {imported_count} 个任务")

        except Exception as e:
            QMessageBox.warning(self, "导入失败", f"导入失败:\n{str(e)}")

    def select_all_tasks(self):
        """全选任务"""
        # 选中所有任务（视觉效果）
        pass

    def delete_completed_tasks(self):
        """删除已完成任务"""
        reply = QMessageBox.question(
            self, "确认删除", "确定要删除所有已完成的任务吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            tasks = self.controller.get_tasks(self.current_list_id)
            for task in tasks:
                if task.is_completed:
                    self.controller.delete_task(task.id)
            self.load_tasks()

    def show_about(self):
        """显示关于对话框"""
        QMessageBox.about(self, "关于 JinSongToDo",
                         "JinSongToDo v1.0\n\n"
                         "一个简洁高效的待办事项管理应用。\n\n"
                         "基于 PySide6 开发")

    def show_shortcuts(self):
        """显示快捷键说明"""
        shortcuts_text = """
快捷键说明:

Ctrl+N - 新建任务
Ctrl+Q - 退出程序
Ctrl+A - 全选任务

在任务列表中:
- 勾选框 - 完成/未完成
- 点击编辑按钮 - 编辑任务
- 点击删除按钮 - 删除任务

拖拽文件到窗口 - 创建文件任务
        """
        QMessageBox.information(self, "快捷键说明", shortcuts_text.strip())

    def closeEvent(self, event):
        """窗口关闭事件 - 最小化到托盘而不是退出"""
        event.ignore()
        self.hide()
        self.tray_icon.showMessage(
            "JinSongToDo",
            "程序已最小化到系统托盘，点击托盘图标可恢复",
            QSystemTrayIcon.Information,
            2000
        )

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 左侧导航栏
        self.sidebar = self.create_sidebar()
        main_layout.addWidget(self.sidebar, 1)

        # 右侧任务区域
        self.task_area = self.create_task_area()
        main_layout.addWidget(self.task_area, 4)

        # 状态栏
        self.statusBar().showMessage("准备就绪")

    def create_sidebar(self):
        sidebar = QWidget()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(180)

        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(8, 16, 8, 8)
        layout.setSpacing(4)

        # 标题
        title = QLabel("JinSongToDo")
        title.setObjectName("sidebarTitle")
        layout.addWidget(title)

        layout.addSpacing(16)

        # 所有任务列表标签
        list_label = QLabel("任务列表")
        list_label.setObjectName("listLabel")
        layout.addWidget(list_label)

        # 列表区域（可滚动）
        self.list_scroll = QScrollArea()
        self.list_scroll.setWidgetResizable(True)
        self.list_scroll.setMaximumHeight(300)
        self.list_scroll.setObjectName("listScroll")

        list_container = QWidget()
        self.list_layout = QVBoxLayout(list_container)
        self.list_layout.setSpacing(2)
        self.list_layout.setContentsMargins(0, 0, 0, 0)

        self.list_scroll.setWidget(list_container)
        layout.addWidget(self.list_scroll)

        # 添加列表按钮
        self.add_list_btn = QPushButton("+ 添加列表")
        self.add_list_btn.setObjectName("addListButton")
        self.add_list_btn.clicked.connect(self.add_new_list)
        layout.addWidget(self.add_list_btn)

        layout.addStretch()

        return sidebar

    def create_task_area(self):
        container = QWidget()
        container.setObjectName("taskArea")

        layout = QVBoxLayout(container)
        layout.setContentsMargins(32, 24, 32, 12)
        layout.setSpacing(0)

        # ========== Header 区域 ==========
        header_widget = QWidget()
        header_widget.setObjectName("headerWidget")
        header_layout = QVBoxLayout(header_widget)
        header_layout.setContentsMargins(0, 0, 0, 20)
        header_layout.setSpacing(8)

        # 日期标签（绿色胶囊）
        from datetime import datetime
        today = datetime.now()
        date_str = today.strftime("%Y年%m月%d日")
        self.date_label = QLabel(date_str)
        self.date_label.setObjectName("dateLabel")
        header_layout.addWidget(self.date_label)

        # 标题行：标题 + 进度条（在右侧）
        title_row_widget = QWidget()
        title_row_layout = QHBoxLayout(title_row_widget)
        title_row_layout.setContentsMargins(0, 0, 0, 0)
        title_row_layout.setSpacing(16)

        # 左侧：标题 + 引言
        left_column = QVBoxLayout()
        left_column.setSpacing(4)

        self.list_title = QLabel("任务清单")
        self.list_title.setObjectName("listTitle")
        left_column.addWidget(self.list_title)

        quotes = [
            "完成的秘诀在于开始",
            "小步前进，大有可为",
            "今日事今日毕",
            "行动是治愈恐惧的良药",
            "一步一个脚印，终将抵达。",
            "专注当下，方能成事。",
        ]
        import random
        self.quote_label = QLabel(random.choice(quotes))
        self.quote_label.setObjectName("quoteLabel")
        left_column.addWidget(self.quote_label)

        title_row_layout.addLayout(left_column, 1)

        # 右侧：进度（百分比在上，进度条在下）
        right_column = QVBoxLayout()
        right_column.setSpacing(4)

        self.progress_text = QLabel("0%")
        self.progress_text.setObjectName("progressText")
        right_column.addWidget(self.progress_text)

        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progressBar")
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(5)
        self.progress_bar.setTextVisible(False)
        right_column.addWidget(self.progress_bar)

        title_row_layout.addLayout(right_column, 1)

        # 置顶按钮（最右边，使用绿色五角星）
        self.pin_btn = QPushButton()
        self.pin_btn.setObjectName("pinButton")
        self.pin_btn.setFixedSize(32, 32)
        self.pin_btn.setCheckable(True)
        self.pin_btn.setStyleSheet("background-color: rgba(90,143,110,0.2); border-radius: 6px;")
        self.pin_btn.setToolTip("置顶窗口")
        self.pin_btn.clicked.connect(self.toggle_pin)

        # 创建绿色五角星图标（缩小30%）
        def create_star_icon(color="#5a8f6e"):
            pixmap = QPixmap(32, 32)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing)
            c = QColor(color)
            painter.setPen(c)
            painter.setBrush(c)
            # 绘制五角星（缩小30%：外径8.4，内径3.5）
            import math
            star_path = QPainterPath()
            for i in range(10):
                r = 8.4 if i % 2 == 0 else 3.5
                angle = math.pi / 2 + i * math.pi / 5
                x = 16 + r * math.cos(angle)
                y = 16 - r * math.sin(angle)
                if i == 0:
                    star_path.moveTo(x, y)
                else:
                    star_path.lineTo(x, y)
            star_path.closeSubpath()
            painter.drawPath(star_path)
            painter.end()
            return QIcon(pixmap)

        self.pin_btn.setIcon(create_star_icon())
        self.pin_btn.setIconSize(QPixmap(32, 32).size())
        title_row_layout.addWidget(self.pin_btn)

        header_layout.addWidget(title_row_widget)
        layout.addWidget(header_widget)

        # ========== Input 区域 ==========
        input_widget = QWidget()
        input_widget.setObjectName("inputWidget")
        input_layout = QHBoxLayout(input_widget)
        input_layout.setContentsMargins(0, 0, 0, 20)
        input_layout.setSpacing(12)

        self.quick_add_input = QLineEdit()
        self.quick_add_input.setPlaceholderText("添加新任务...")
        self.quick_add_input.setObjectName("quickAddInput")
        self.quick_add_input.returnPressed.connect(self.quick_add_task)
        input_layout.addWidget(self.quick_add_input, 1)

        self.add_btn = QPushButton("+")
        self.add_btn.setObjectName("addButton")
        self.add_btn.setFixedSize(46, 46)
        self.add_btn.clicked.connect(self.quick_add_task)
        input_layout.addWidget(self.add_btn)

        layout.addWidget(input_widget)

        # ========== Filter Tabs 区域 ==========
        filter_widget = QWidget()
        filter_widget.setObjectName("filterWidget")
        filter_layout = QHBoxLayout(filter_widget)
        filter_layout.setContentsMargins(0, 0, 0, 20)
        filter_layout.setSpacing(4)

        self.filter_buttons = []
        filter_options = [("全部", "all"), ("进行中", "active"), ("已完成", "completed")]

        from PySide6.QtWidgets import QButtonGroup
        self.filter_group = QButtonGroup()

        for text, key in filter_options:
            btn = QPushButton(text)
            btn.setObjectName("filterButton")
            btn.setCheckable(True)
            btn.setChecked(key == "all")
            self.filter_group.addButton(btn)
            self.filter_buttons.append((btn, key))
            filter_layout.addWidget(btn)

        filter_layout.addStretch()

        # 连接筛选按钮
        self.filter_group.buttonClicked.connect(self.on_filter_changed)

        layout.addWidget(filter_widget)

        # ========== 任务列表容器 ==========
        self.task_container = QWidget()
        task_container_layout = QVBoxLayout(self.task_container)
        task_container_layout.setContentsMargins(0, 0, 0, 0)
        task_container_layout.setSpacing(0)

        # 任务列表
        self.task_scroll = QScrollArea()
        self.task_scroll.setWidgetResizable(True)
        self.task_scroll.setObjectName("taskScroll")
        self.task_scroll.setAcceptDrops(True)

        self.task_widget = QWidget()

        # 任务列表布局
        self.task_list_layout = QVBoxLayout(self.task_widget)
        self.task_list_layout.setSpacing(8)
        self.task_list_layout.setContentsMargins(0, 0, 0, 0)
        self.task_list_layout.setAlignment(Qt.AlignTop)

        self.task_scroll.setWidget(self.task_widget)
        task_container_layout.addWidget(self.task_scroll)

        # 安装事件过滤器捕获拖拽事件
        self.task_scroll.viewport().installEventFilter(self)
        self.task_scroll.viewport().setAcceptDrops(True)
        self.task_scroll.installEventFilter(self)
        self.task_scroll.setAcceptDrops(True)
        self.task_container.installEventFilter(self)
        self.task_container.setAcceptDrops(True)

        # 拖拽提示区域
        self.drop_zone = QWidget()
        self.drop_zone.setObjectName("dropZone")
        drop_layout = QVBoxLayout(self.drop_zone)
        drop_layout.setAlignment(Qt.AlignCenter)
        drop_icon = QLabel("📎")
        drop_icon.setObjectName("dropIcon")
        drop_icon.setAlignment(Qt.AlignCenter)
        drop_layout.addWidget(drop_icon)
        drop_text = QLabel("拖放文件到此处创建任务")
        drop_text.setObjectName("dropText")
        drop_text.setAlignment(Qt.AlignCenter)
        drop_layout.addWidget(drop_text)
        self.drop_zone.hide()
        task_container_layout.addWidget(self.drop_zone)

        layout.addWidget(self.task_container, 1)

        # ========== Footer 区域 ==========
        footer_widget = QWidget()
        footer_widget.setObjectName("footerWidget")
        footer_layout = QHBoxLayout(footer_widget)
        footer_layout.setContentsMargins(0, 16, 0, 0)
        footer_layout.setSpacing(8)

        self.clear_btn = QPushButton("清除已完成")
        self.clear_btn.setObjectName("clearButton")
        self.clear_btn.clicked.connect(self.clear_completed)
        footer_layout.addWidget(self.clear_btn)
        footer_layout.addStretch()

        self.clear_count_label = QLabel("")
        self.clear_count_label.setObjectName("clearCountLabel")
        footer_layout.addWidget(self.clear_count_label)

        layout.addWidget(footer_widget)

        return container

    def apply_theme(self):
        stylesheet = self.theme_manager.generate_stylesheet()
        self.setStyleSheet(stylesheet)

    def change_theme(self, theme_key):
        self.theme_manager.set_theme(theme_key)
        self.apply_theme()
        # 更新所有卡片样式
        self._update_all_card_styles()

    def _update_all_card_styles(self):
        for i in range(self.task_list_layout.count()):
            item = self.task_list_layout.itemAt(i)
            if item and item.widget() and hasattr(item.widget(), 'update_style'):
                item.widget().update_style()

    def on_list_selected(self, list_id, list_name):
        # 先更新主题样式
        self.apply_theme()

        self.current_list_id = list_id
        self.current_list_name = list_name
        self.list_title.setText(list_name)

        # 更新列表按钮状态
        for btn in self.list_buttons:
            btn.setChecked(False)

        # 找到并选中对应的按钮
        lists = self.controller.get_all_lists()
        for i, lst in enumerate(lists):
            if lst.id == list_id and i < len(self.list_buttons):
                self.list_buttons[i].setChecked(True)
                break

        self.load_tasks()

    def load_lists(self):
        # 清空现有列表按钮
        while self.list_layout.count():
            child = self.list_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # 加载用户列表
        lists = self.controller.get_all_lists()
        self.list_buttons = []

        for lst in lists:
            btn = QPushButton(f"{lst.icon} {lst.name}")
            btn.setObjectName("listButton")
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, lid=lst.id, lname=lst.name:
                              self.on_list_selected(lid, lname))
            # 右键菜单 - 编辑/删除列表
            btn.setContextMenuPolicy(Qt.CustomContextMenu)
            btn.customContextMenuRequested.connect(lambda pos, lid=lst.id, b=btn: self.show_list_menu(pos, lid, b))
            self.list_layout.addWidget(btn)
            self.list_buttons.append(btn)

    def show_list_menu(self, pos, list_id, btn):
        # 检查是否是系统列表
        lst = self.controller.db.get_list(list_id)
        if lst and lst.is_system:
            return  # 系统列表不显示菜单

        menu = QMenu()
        edit_action = QAction("编辑列表", menu)
        edit_action.triggered.connect(lambda: self.edit_list(list_id))
        delete_action = QAction("删除列表", menu)
        delete_action.triggered.connect(lambda: self.delete_list(list_id))
        menu.addAction(edit_action)
        menu.addAction(delete_action)
        menu.exec(btn.mapToGlobal(pos))

    def edit_list(self, list_id):
        lst = self.controller.db.get_list(list_id)
        if not lst or lst.is_system:
            return  # 不能编辑系统列表

        dialog = QDialog(self)
        dialog.setWindowTitle("编辑列表")
        dialog.setMinimumWidth(350)

        layout = QVBoxLayout(dialog)

        # 名称
        name_layout = QHBoxLayout()
        name_label = QLabel("名称:")
        name_label.setFixedWidth(40)
        name_layout.addWidget(name_label)
        name_input = QLineEdit()
        name_input.setText(lst.name)
        name_input.setFixedWidth(180)
        name_layout.addWidget(name_input)
        name_layout.addStretch()
        layout.addLayout(name_layout)

        # 图标选择 - 使用下拉框
        icon_layout = QHBoxLayout()
        icon_label = QLabel("图标:")
        icon_label.setFixedWidth(40)
        icon_layout.addWidget(icon_label)
        icon_combo = QComboBox()
        icon_combo.setEditable(True)
        icon_combo.setFixedWidth(180)
        emojis = ["📋", "📅", "📆", "⭐", "💼", "🏠", "📚", "🎯", "💡", "🔥", "❤️", "🎉",
                  "✅", "📌", "🔔", "💬", "📝", "🎨", "🏆", "🌟", "💪", "🚀", "🎮", "🎬"]
        for emoji in emojis:
            icon_combo.addItem(emoji)
        icon_combo.setCurrentText(lst.icon)
        icon_layout.addWidget(icon_combo)
        icon_layout.addStretch()
        layout.addLayout(icon_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(lambda: self.save_list_edit(list_id, name_input.text(), icon_combo.currentText(), dialog))
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        dialog.exec()

    def select_emoji(self, emoji, input_field):
        input_field.setText(emoji)

    def save_list_edit(self, list_id, name, icon, dialog):
        if name.strip():
            self.controller.update_list(list_id, name, icon=icon if icon else "📋")
            self.load_lists()
            dialog.accept()

    def delete_list(self, list_id):
        lst = self.controller.db.get_list(list_id)
        if lst and lst.is_system:
            QMessageBox.information(self, "提示", "系统列表不能删除")
            return

        reply = QMessageBox.question(
            self, "确认删除", "确定要删除这个列表吗？列表下的所有任务也会被删除。",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.controller.delete_list(list_id)
            self.load_lists()
            if self.current_list_id == list_id:
                self.on_list_selected(1, "我的日程")

    def load_tasks(self):
        # 清空现有任务
        while self.task_list_layout.count():
            child = self.task_list_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # 加载任务
        all_tasks = self.controller.get_tasks(self.current_list_id, None)

        # 应用筛选
        if self.current_filter == "active":
            tasks = [t for t in all_tasks if not t.is_completed]
        elif self.current_filter == "completed":
            tasks = [t for t in all_tasks if t.is_completed]
        else:
            tasks = all_tasks

        for task in tasks:
            card = TaskCard(task, self.theme_manager)
            card.delete_signal.connect(self.delete_task)
            card.toggle_signal.connect(self.toggle_task)
            self.task_list_layout.insertWidget(0, card)

        # 更新所有任务卡片的样式
        def update_cards():
            for i in range(self.task_list_layout.count()):
                item = self.task_list_layout.itemAt(i)
                if item and item.widget() and hasattr(item.widget(), 'update_style'):
                    item.widget().update_style()

        update_cards()
        # 延迟再次更新确保UI刷新
        QTimer.singleShot(10, update_cards)

        # 确保滚动区域内容顶部对齐
        self.task_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)

        # 根据任务数量决定是否显示拖放提示
        if len(tasks) == 0:
            self.drop_zone.show()
        else:
            self.drop_zone.hide()

        # 更新进度
        total = len(all_tasks)
        completed = sum(1 for t in all_tasks if t.is_completed)
        self.update_progress(completed, total)

        # 更新状态栏
        self.update_status()

    def on_search(self, text):
        self.load_tasks()

    def quick_add_task(self):
        title = self.quick_add_input.text().strip()
        if title:
            self.controller.create_task(
                title=title,
                list_id=self.current_list_id if self.current_list_id > 0 else 0,
            )
            self.quick_add_input.clear()
            self.load_tasks()

    def eventFilter(self, obj, event):
        """事件过滤器 - 处理视口中的拖拽事件"""
        if obj in (self.task_scroll, self.task_scroll.viewport(), self.task_container):
            etype = event.type()
            if etype == QEvent.DragEnter:
                if event.mimeData().hasUrls():
                    event.acceptProposedAction()
                    self.drop_zone.show()
                    return True
            elif etype == QEvent.DragLeave:
                self.drop_zone.hide()
                return True
            elif etype == QEvent.Drop:
                mime_data = event.mimeData()
                if mime_data.hasUrls():
                    for url in mime_data.urls():
                        file_path = url.toLocalFile()
                        import os
                        if os.path.isfile(file_path):
                            self.create_task_from_file(file_path)
                    event.acceptProposedAction()
                    self.drop_zone.hide()
                    return True
        return super().eventFilter(obj, event)

    def create_task_from_file(self, file_path):
        import os
        # 获取文件名（去掉扩展名）
        filename = os.path.basename(file_path)
        name = os.path.splitext(filename)[0]
        # 文件路径作为备注
        description = f"📎 文件路径: {file_path}"
        self.controller.create_task(
            title=name,
            description=description,
            list_id=self.current_list_id if self.current_list_id > 0 else 0,
        )
        self.load_tasks()
        self.statusBar().showMessage(f"已添加文件任务: {name}", 3000)

    # ==================== 任务拖拽排序 ====================

    def task_scroll_dragMove(self, event):
        """任务滚动区拖拽移动"""
        if event.mimeData().hasText() and not event.mimeData().text().startswith('file:///'):
            pos = event.pos()
            self.update_drop_indicator(pos)
            event.acceptProposedAction()
        else:
            event.ignore()

    def task_scroll_dragLeave(self, event):
        """任务滚动区拖拽离开"""
        self.drop_indicator.hide()

    def task_scroll_drop(self, event):
        """任务滚动区放下"""
        self.drop_indicator.hide()
        mime = event.mimeData()
        if mime.hasText():
            text = mime.text()
            if text.startswith('file:///'):
                # 文件拖放
                for url in mime.urls():
                    file_path = url.toLocalFile()
                    import os
                    if os.path.isfile(file_path):
                        self.create_task_from_file(file_path)
            else:
                pass
        event.acceptProposedAction()
        event.setAccepted(True)

    def task_widget_dragMove(self, event):
        """任务容器拖拽移动"""
        if event.mimeData().hasText() and not event.mimeData().text().startswith('file:///'):
            pos = self.task_widget.mapFromParent(event.pos())
            self.update_drop_indicator(pos)
            event.acceptProposedAction()
        else:
            event.ignore()

    def task_widget_dragLeave(self, event):
        """任务容器拖拽离开"""
        self.drop_indicator.hide()

    def task_widget_drop(self, event):
        """任务容器放下"""
        mime = event.mimeData()
        if mime.hasText():
            text = mime.text()
            if text.startswith('file:///'):
                for url in mime.urls():
                    file_path = url.toLocalFile()
                    import os
                    if os.path.isfile(file_path):
                        self.create_task_from_file(file_path)
            else:
                try:
                    task_id = int(text)
                    pos = self.task_widget.mapFromParent(event.pos())
                    drop_index = self.get_drop_index(pos)
                    self.reorder_task(task_id, drop_index)
                except:
                    pass
        event.acceptProposedAction()
        event.setAccepted(True)

    def add_task(self):
        dialog = TaskEditDialog(parent=self)
        if dialog.exec():
            data = dialog.get_task_data()
            self.controller.create_task(
                title=data['title'],
                description=data['description'],
                is_important=data['is_important'],
                due_date=data['due_date'],
                due_time=data['due_time'],
                list_id=self.current_list_id if self.current_list_id > 0 else 0,
            )
            self.load_tasks()

    def edit_task(self, task):
        dialog = TaskEditDialog(task, parent=self)
        if dialog.exec():
            data = dialog.get_task_data()
            task.title = data['title']
            task.description = data['description']
            task.is_important = data['is_important']
            task.due_date = data['due_date']
            task.due_time = data['due_time']
            self.controller.update_task(task)
            self.load_tasks()

    def delete_task(self, task_id):
        reply = QMessageBox.question(
            self, "确认删除", "确定要删除这个任务吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.controller.delete_task(task_id)
            self.load_tasks()

    def toggle_task(self, task_id):
        self.controller.toggle_completed(task_id)
        self.load_tasks()

    def add_new_list(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("添加列表")
        dialog.setMinimumWidth(300)

        layout = QVBoxLayout(dialog)

        # 名称
        name_layout = QHBoxLayout()
        name_label = QLabel("名称:")
        name_label.setFixedWidth(40)
        name_layout.addWidget(name_label)
        name_input = QLineEdit()
        name_input.setPlaceholderText("输入列表名称")
        name_input.setFixedWidth(180)
        name_layout.addWidget(name_input)
        name_layout.addStretch()
        layout.addLayout(name_layout)

        # 图标选择 - 使用下拉框
        icon_layout = QHBoxLayout()
        icon_label = QLabel("图标:")
        icon_label.setFixedWidth(40)
        icon_layout.addWidget(icon_label)
        icon_combo = QComboBox()
        icon_combo.setEditable(True)
        icon_combo.setFixedWidth(180)
        emojis = ["📋", "📅", "📆", "⭐", "💼", "🏠", "📚", "🎯", "💡", "🔥", "❤️", "🎉",
                  "✅", "📌", "🔔", "💬", "📝", "🎨", "🏆", "🌟", "💪", "🚀", "🎮", "🎬"]
        for emoji in emojis:
            icon_combo.addItem(emoji)
        icon_combo.setCurrentText("📋")
        icon_layout.addWidget(icon_combo)
        icon_layout.addStretch()
        layout.addLayout(icon_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(lambda: self.save_new_list(name_input.text(), icon_combo.currentText(), dialog))
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        dialog.exec()

    def save_new_list(self, name, icon, dialog):
        if name.strip():
            self.controller.create_list(name, icon=icon if icon else "📋")
            self.load_lists()
            dialog.accept()

    def update_status(self):
        completed, total = self.controller.get_today_stats()
        self.statusBar().showMessage(f"今日已完成 {completed}/{total}")

    def on_filter_changed(self, button):
        """筛选按钮切换"""
        for btn, key in self.filter_buttons:
            if btn == button:
                self.current_filter = key
                btn.setChecked(True)
            else:
                btn.setChecked(False)
        self.load_tasks()

    def clear_completed(self):
        """清除已完成的任务"""
        self.controller.delete_completed_tasks(self.current_list_id)
        self.load_tasks()

    def update_progress(self, completed, total):
        """更新进度显示"""
        if total == 0:
            percent = 0
        else:
            percent = int(completed / total * 100)
        self.progress_bar.setValue(percent)
        self.progress_text.setText(f"{percent}%")

        # 更新清除按钮计数
        completed_count = self.controller.get_completed_count(self.current_list_id)
        if completed_count > 0:
            self.clear_count_label.setText(f"({completed_count}) >")
            self.clear_btn.setVisible(True)
        else:
            self.clear_count_label.setText("")
            self.clear_btn.setVisible(False)